"""XLSX generation via openpyxl. Markdown table OR metadata['sheets'] (list of sheets)."""
import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _parse_md_table(content: str) -> List[List[str]]:
    lines = content.split("\n")
    table = []
    in_table = False
    for line in lines:
        s = line.strip()
        if s.startswith("|") and "|" in s[1:]:
            cells = [c.strip() for c in s.strip("|").split("|")]
            if not in_table:
                in_table = True
                table.append(cells)
            else:
                if all(set(c) <= set("-: ") for c in cells):
                    continue
                table.append(cells)
        elif in_table:
            break
    return table


def _write_sheet(ws, rows: List[List[str]]):
    if not rows:
        return
    for ri, row in enumerate(rows, start=1):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            if ri == 1:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal="left", vertical="center")


def render(
    title: str,
    content: str,
    output_path: Path,
    metadata: Dict[str, Any] = None,
) -> Path:
    metadata = metadata or {}
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets_meta = metadata.get("sheets")
    if sheets_meta and isinstance(sheets_meta, list):
        for sm in sheets_meta:
            ws = wb.create_sheet(title=sm.get("name", "Sheet")[:31])
            _write_sheet(ws, sm.get("rows", []))
    else:
        rows = _parse_md_table(content)
        if not rows:
            rows = [["content"], [content]]
        ws = wb.create_sheet(title=(title or "Sheet")[:31])
        _write_sheet(ws, rows)

    wb.save(str(output_path))
    return output_path
